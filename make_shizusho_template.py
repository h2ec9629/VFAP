# -*- coding: utf-8 -*-
"""
指図書ブロック テンプレート生成スクリプト
────────────────────────────────────────────────────────────
貼り付け済み指示書ファイルから 38:56 行ブロック
（罫線・結合セル・行高・静的ラベル）を抽出し
shizusho_template.xlsx を生成する。

・SGT.xlsm / VFDB.xlsm への外部リンク式や実績値セルは残さない
　（値は印刷時に _fill_jisseki が vfdb.json / kakou_kiroku.json から書き込む）
・ローカル数式（B39/B44/F44）だけ数式のまま保持
・レイアウトを変えたいときは SRC_FILE を差し替えて再実行するだけ

使い方: python make_shizusho_template.py
"""
from pathlib import Path
from copy import copy
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula

SRC_FILE = Path(__file__).resolve().parent / "input" / "樹脂指示書" / "260526_灯具カバー（526）.xlsx"
DST_FILE = Path(__file__).resolve().parent / "shizusho_template.xlsx"

ROW_MIN, ROW_MAX = 38, 56
COL_MIN, COL_MAX = 1, 14   # A〜N

# 値を残す静的ラベルセル
KEEP_VALUES = {
    "A38", "B38", "C38", "E38", "F38", "G38", "H38", "I38",
    "I39", "I40", "I41", "I42", "I43", "I44", "I45",
    "A40", "B40", "A42", "B42", "C41", "C43",
    "A44", "C44", "D44", "E44", "G44", "H44",
    "I46", "I47", "I48", "I49", "I50", "I51", "I52", "I53", "I54",
    "A55", "I55",
    # 不良カウント欄の既定値 0（印刷時に実績があれば上書きされる）
    "J40", "J41", "J42", "J43", "J44",
}

# ローカル数式（外部リンクなし）は数式のまま残す
KEEP_LOCAL_FORMULAS = {
    "B39": "=D3",
}

# 配列数式（CSE）で残すセル ※普通の数式だとLibreOfficeで#VALUE!になる
KEEP_ARRAY_FORMULAS = {
    "B44": '=IF(SUM(--(ISTEXT(B45:B55)*(B45:B55<>"")))=0,"計測Lot","計測Lot・加工数")',
    "F44": '=IF(SUM(--(ISTEXT(B45:B55)*(B45:B55<>"")))=0,"記入不要","使用Lot")',
}


def main():
    src_wb = openpyxl.load_workbook(SRC_FILE)
    src_ws = src_wb.worksheets[0]

    dst_wb = openpyxl.Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "指図書ブロック"

    for row in src_ws.iter_rows(min_row=ROW_MIN, max_row=ROW_MAX,
                                min_col=COL_MIN, max_col=COL_MAX):
        for c in row:
            d = dst_ws.cell(c.row, c.column)
            # スタイル（罫線・フォント・塗り・配置・表示形式）
            if c.has_style:
                d.font          = copy(c.font)
                d.border        = copy(c.border)
                d.fill          = copy(c.fill)
                d.alignment     = copy(c.alignment)
                d.number_format = c.number_format
                d.protection    = copy(c.protection)
            # 値：静的ラベルとローカル数式だけ
            coord = c.coordinate
            if coord in KEEP_LOCAL_FORMULAS:
                d.value = KEEP_LOCAL_FORMULAS[coord]
            elif coord in KEEP_ARRAY_FORMULAS:
                d.value = ArrayFormula(coord, KEEP_ARRAY_FORMULAS[coord])
            elif coord in KEEP_VALUES:
                d.value = c.value

    # 結合セル（38:56 内のみ）
    for m in src_ws.merged_cells.ranges:
        if ROW_MIN <= m.min_row and m.max_row <= ROW_MAX:
            dst_ws.merge_cells(str(m))

    # 行高
    for r in range(ROW_MIN, ROW_MAX + 1):
        h = src_ws.row_dimensions[r].height
        if h is not None:
            dst_ws.row_dimensions[r].height = h

    dst_wb.save(DST_FILE)
    print(f"OK: {DST_FILE.name} を生成しました")
    print(f"  結合セル: {len([m for m in dst_ws.merged_cells.ranges])} 個")


if __name__ == "__main__":
    main()
