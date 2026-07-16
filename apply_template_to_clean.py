# -*- coding: utf-8 -*-
"""
既存クリーン版への指図書ブロック一括適用スクリプト
────────────────────────────────────────────────────────────
樹脂指示書_クリーン 内の全xlsxに shizusho_template.xlsx の
ブロック（38:56行）を上書き適用する。

・app.py の paste_shizusho_from_template をそのまま抽出して使う
　（ロジックの二重管理をしない）
・ページ数はファイル内の各57行チャンクに内容があるかで自動判定
・実行前にバックアップ推奨（例: _backup_クリーン_YYYYMMDD.zip）

使い方: python apply_template_to_clean.py
"""
import ast
from pathlib import Path
import openpyxl

BASE = Path(__file__).resolve().parent
CLEAN_DIR = BASE / "樹脂指示書_クリーン"


def load_paste_fn():
    """app.py から paste_shizusho_from_template を抽出して返す"""
    src = (BASE / "app.py").read_text(encoding="utf-8")
    tree = ast.parse(src)
    fn_src = next(ast.get_source_segment(src, n) for n in tree.body
                  if isinstance(n, ast.FunctionDef)
                  and n.name == "paste_shizusho_from_template")
    g = {"Path": Path, "openpyxl": openpyxl, "CLEAN_DIR": CLEAN_DIR,
         "__file__": str(BASE / "app.py")}
    exec(compile(fn_src, "app.py:paste_shizusho_from_template", "exec"), g)
    return g["paste_shizusho_from_template"]


def detect_pages(path: Path) -> list:
    """57行チャンクごとに帳票ヘッダー（品目番号B3等）の有無でページ数を判定"""
    ws = openpyxl.load_workbook(path, read_only=True).active
    pages = []
    max_page = max(1, (ws.max_row + 56) // 57)
    for n in range(1, max_page + 1):
        off = (n - 1) * 57
        if any(ws.cell(off + r, 2).value not in (None, "") for r in (3, 5, 7)):
            pages.append(n)
    return pages or [1]


def main():
    paste = load_paste_fn()
    targets = []
    for f in sorted(CLEAN_DIR.glob("*.xlsx")):
        for pg in detect_pages(f):
            targets.append((str(f), pg))
    print(f"対象: {len(set(t[0] for t in targets))}ファイル / {len(targets)}ページ")
    ok, msg = paste(targets)
    print("OK" if ok else "NG", "-", msg)


if __name__ == "__main__":
    main()
