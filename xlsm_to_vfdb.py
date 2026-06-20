#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
xlsm_to_vfdb.py ── VFDB.xlsm(シート VFDB2) → vfdb.json 変換スクリプト

品目マスタ(VFDB.xlsm)を編集したら、これ1発で配信用 vfdb.json を再生成する。
既定は DRY-RUN(差分プレビューのみ、書き込まない)。--apply で実書き込み(.bak退避つき)。

使い方(Windows / リポジトリ直下で):
    python xlsm_to_vfdb.py            # 差分プレビューだけ(安全)
    python xlsm_to_vfdb.py --apply    # vfdb.json を上書き(.bak_日時 作成)

入出力パスは既定で OneDrive の実レイアウトを見にいく。--in / --out で上書き可。

列ごとの型ルールは現行 vfdb.json のクセに合わせて確定済み(下記 STR_COLS / NUM_COLS)。
VFDB2 の列構成(順番含む)を変えたら、この json も自動でその順になる。
"""
import os, sys, json, argparse, datetime, shutil
from pathlib import Path

SHEET = "VFDB2"

# ── 列ごとの値変換ルール(現行 vfdb.json の型クセに合わせて確定) ──
#   既定(RAW): Excel の値そのまま。整数の float(36.0)は int(36)へ。
#   STR     : 文字列化。全角文字・全角スペース・記号は原文のまま保持。
#   NUM     : 符号付き/全角数字の文字列("+0.5""591.0")を数値化。空白のみは null。
STR_COLS = {"下限公差", "取数", "梱包方法", "ラベル", "切断寸法"}
NUM_COLS = {"上限公差", "下限", "センター", "上限"}


def t_raw(v):
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def t_str(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v)


def t_num(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace("＋", "+").replace("－", "-")
    try:
        f = float(s)
        return int(f) if f.is_integer() else f
    except ValueError:
        return str(v)


def convert_cell(col, v):
    if col in STR_COLS:
        return t_str(v)
    if col in NUM_COLS:
        return t_num(v)
    return t_raw(v)


def onedrive_work():
    up = os.environ.get("USERPROFILE", "")
    if not up:
        return None
    cand = Path(up) / "OneDrive" / "work"
    return cand if cand.exists() else None


def default_in():
    ow = onedrive_work()
    return (ow / "SGT_cloud" / "master" / "VFDB.xlsm") if ow else Path("VFDB.xlsm")


def default_out():
    ow = onedrive_work()
    return (ow / "VFAP" / "vfdb.json") if ow else Path("vfdb.json")


def read_xlsm(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"[NG] シート '{SHEET}' が見つからへん: {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        sys.exit("[NG] VFDB2 が空")
    hdr = list(rows[0])
    recs = []
    for r in rows[1:]:
        if not r[0]:                       # 品目コード空の行はスキップ
            continue
        recs.append({hdr[i]: convert_cell(hdr[i], r[i]) for i in range(len(hdr))})
    return hdr, recs


def dump_json(recs):
    s = json.dumps(recs, ensure_ascii=False, indent=2)
    return s.replace("\n", "\r\n")         # 現行ファイルに合わせ CRLF / 末尾改行なし


def load_existing(path):
    p = Path(path)
    if not p.exists():
        return None
    try:
        return json.loads(p.read_text("utf-8"))
    except Exception as e:
        print(f"[警告] 既存 vfdb.json の読込に失敗(差分は出せません): {e}")
        return None


def show_diff(old, new):
    if old is None:
        print("既存 vfdb.json 無し → 新規作成扱い")
        return
    ox = {str(r.get("品目コード")): r for r in old}
    nx = {str(r.get("品目コード")): r for r in new}
    added   = sorted(set(nx) - set(ox))
    removed = sorted(set(ox) - set(nx))
    changed = []
    for code in sorted(set(ox) & set(nx)):
        o, n = ox[code], nx[code]
        cells = [(k, o.get(k), n.get(k)) for k in n if o.get(k) != n.get(k)]
        if cells:
            changed.append((code, n.get("品名"), cells))
    print("\n=== 差分(既存 vfdb.json → xlsm から再生成) ===")
    print(f"追加 {len(added)}品目 / 削除 {len(removed)}品目 / 変更 {len(changed)}品目")
    for c in added:
        print(f"  + {c}")
    for c in removed:
        print(f"  - {c}")
    for code, name, cells in changed:
        print(f"  ~ {code}  {name}")
        for k, ov, nv in cells:
            print(f"        {k}: {ov!r} -> {nv!r}")
    if not (added or removed or changed):
        print("  差分なし(完全一致)")


def main():
    ap = argparse.ArgumentParser(description="VFDB.xlsm(VFDB2) → vfdb.json 変換")
    ap.add_argument("--in",  dest="src", default=str(default_in()))
    ap.add_argument("--out", dest="dst", default=str(default_out()))
    ap.add_argument("--apply", action="store_true",
                    help="実際に書き込む(既定は DRY-RUN で差分表示のみ)")
    a = ap.parse_args()

    src, dst = Path(a.src), Path(a.dst)
    print(f"IN : {src}")
    print(f"OUT: {dst}")
    if not src.exists():
        sys.exit(f"[NG] 入力ファイルが無い: {src}")

    hdr, recs = read_xlsm(src)
    print(f"読込: {len(recs)}品目 / {len(hdr)}列")

    show_diff(load_existing(dst), recs)

    out = dump_json(recs)
    if not a.apply:
        print("\n[DRY-RUN] 書き込みしてへん。確定するなら --apply を付けて再実行。")
        return

    if dst.exists():
        bak = dst.with_name(dst.name + ".bak_" + datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
        shutil.copy2(dst, bak)
        print(f"バックアップ: {bak}")
    tmp = dst.with_name(dst.name + ".tmp")
    tmp.write_bytes(out.encode("utf-8"))
    os.replace(tmp, dst)
    print(f"[OK] 書き込み完了: {dst} ({len(recs)}品目)")


if __name__ == "__main__":
    main()
